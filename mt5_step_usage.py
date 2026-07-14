#!/usr/bin/env python3
"""Build an Excel report showing which optimization grid steps survivors used."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from mt5_set_files import (
    choose_base_set,
    canonical_set_value,
    discover_set_files,
    ParamGrid,
    parse_set_file,
    parse_set_grid,
    sanitize,
)

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from mt5_paths import DEFAULT_BEST_DIR, resolve_set_dir

DEFAULT_TERMINAL = Path(r"C:\Program Files\MetaTrader 5\terminal64.exe")
DEFAULT_OUTPUT = Path("reports") / "step_usage.xlsx"
METRIC_COLUMNS = [
    "validation_score",
    "validation_recovery",
    "validation_sharpe",
    "validation_cagr_pct",
    "realticks_equity_dd_pct",
    "ohlc_equity_dd_pct",
    "scaled_risk",
]
SURVIVOR_ID_COLUMNS = ["base_set", "profile", "symbol", "timeframe", "pass_id"]
SURVIVOR_COLUMNS = SURVIVOR_ID_COLUMNS + METRIC_COLUMNS
COUNT_COLUMNS = ["base_set", "param", "step_value", "count", "distinct_symbols", "pct"]


@dataclass(frozen=True)
class SurvivorRow:
    base_set: str
    profile: str
    symbol: str
    timeframe: str
    pass_id: str
    metrics: dict[str, str]
    values: dict[str, str]


@dataclass(frozen=True)
class CountRow:
    base_set: str
    param: str
    step_value: str
    count: int
    distinct_symbols: int
    pct: float


def _metrics(row: dict[str, str]) -> dict[str, str]:
    return {name: row.get(name, "") for name in METRIC_COLUMNS}


def _survivor_row(survivor: SurvivorRow, params: list[str]) -> list[Any]:
    return [
        survivor.base_set,
        survivor.profile,
        survivor.symbol,
        survivor.timeframe,
        survivor.pass_id,
        *[survivor.metrics.get(name, "") for name in METRIC_COLUMNS],
        *[survivor.values.get(param, "") for param in params],
    ]


def _count_row(row: CountRow) -> list[Any]:
    return [row.base_set, row.param, row.step_value, row.count, row.distinct_symbols, row.pct]


def _count_sort_key(row: CountRow) -> tuple[str, str, int, str]:
    return (row.base_set, row.param, row.count, row.step_value)


def _winning_set_path(row: dict[str, str], best_dir: Path) -> Path | None:
    stem = sanitize(f"{row.get('symbol', '')}_{row.get('timeframe', '')}_{row.get('profile', '')}_pass{row.get('pass_id', '')}")
    preferred = best_dir / "sets" / f"{stem}.set"
    if preferred.is_file():
        return preferred

    fallback = row.get("set_file", "").strip()
    if fallback:
        fallback_path = Path(fallback).expanduser()
        if fallback_path.is_file():
            return fallback_path
    return None


def load_survivors(
    *,
    best_dir: Path,
    set_dir: Path,
    set_index: dict[str, Path],
) -> tuple[list[SurvivorRow], dict[str, dict[str, ParamGrid]], list[str]]:
    survivors_csv = best_dir / "best_survivors.csv"
    if not survivors_csv.is_file():
        return [], {}, [f"Missing survivors CSV: {survivors_csv}"]

    grids_by_set: dict[str, dict[str, ParamGrid]] = {}
    failed_sets: set[str] = set()
    survivors: list[SurvivorRow] = []
    warnings: list[str] = []

    with survivors_csv.open(newline="", encoding="utf-8") as f:
        for csv_row in csv.DictReader(f):
            source_xml = csv_row.get("source_xml", "").strip()
            if not source_xml:
                warnings.append(f"Skipping row without source_xml: pass={csv_row.get('pass_id', '')}")
                continue

            try:
                base_set_path = choose_base_set(report_stem=Path(source_xml).stem, set_index=set_index)
            except FileNotFoundError as exc:
                warnings.append(str(exc))
                continue

            try:
                base_set = base_set_path.relative_to(set_dir).as_posix()
            except ValueError:
                base_set = base_set_path.as_posix()

            if base_set in failed_sets:
                continue
            if base_set in grids_by_set:
                grids = grids_by_set[base_set]
            else:
                try:
                    grids = parse_set_grid(base_set_path)
                except ValueError as exc:
                    warnings.append(f"Skipping malformed grid {base_set}: {exc}")
                    failed_sets.add(base_set)
                    continue
                grids_by_set[base_set] = grids

            winning_set = _winning_set_path(csv_row, best_dir)
            if winning_set is None:
                warnings.append(
                    f"Missing winning .set for {csv_row.get('symbol', '')} "
                    f"{csv_row.get('timeframe', '')} {csv_row.get('profile', '')} "
                    f"pass={csv_row.get('pass_id', '')}"
                )
                continue

            chosen = parse_set_file(winning_set)
            missing = [name for name in grids if name not in chosen]
            if missing:
                warnings.append(
                    f"Winning .set missing grid params {base_set} "
                    f"pass={csv_row.get('pass_id', '')}: {', '.join(sorted(missing))}"
                )
            values = {name: canonical_set_value(chosen[name]) for name in grids if name in chosen}
            survivors.append(
                SurvivorRow(
                    base_set=base_set,
                    profile=csv_row.get("profile", ""),
                    symbol=csv_row.get("symbol", ""),
                    timeframe=csv_row.get("timeframe", ""),
                    pass_id=csv_row.get("pass_id", ""),
                    metrics=_metrics(csv_row),
                    values=values,
                )
            )

    return survivors, grids_by_set, warnings


def aggregate_counts(
    survivors: list[SurvivorRow],
    grids_by_set: dict[str, dict[str, ParamGrid]],
) -> tuple[list[CountRow], list[str]]:
    by_set_param: dict[tuple[str, str], list[SurvivorRow]] = defaultdict(list)
    warnings: list[str] = []
    for survivor in survivors:
        for param in grids_by_set.get(survivor.base_set, {}):
            by_set_param[(survivor.base_set, param)].append(survivor)

    rows: list[CountRow] = []
    for base_set, grids in sorted(grids_by_set.items()):
        for param, grid in grids.items():
            scoped = by_set_param.get((base_set, param), [])
            counts: Counter[str] = Counter()
            symbols_by_value: dict[str, set[str]] = defaultdict(set)
            for survivor in scoped:
                value = survivor.values.get(param)
                if value is None:
                    continue
                if value not in grid.steps:
                    warnings.append(f"Off-grid value {base_set} {param}={value} pass={survivor.pass_id}")
                counts[value] += 1
                symbols_by_value[value].add(survivor.symbol)

            total = sum(counts.values())
            for step in grid.steps:
                count = counts[step]
                rows.append(
                    CountRow(
                        base_set=base_set,
                        param=param,
                        step_value=step,
                        count=count,
                        distinct_symbols=len(symbols_by_value[step]),
                        pct=(count / total) if total else 0.0,
                    )
                )
            for value in sorted(set(counts) - set(grid.steps)):
                count = counts[value]
                rows.append(
                    CountRow(
                        base_set=base_set,
                        param=param,
                        step_value=f"off_grid:{value}",
                        count=count,
                        distinct_symbols=len(symbols_by_value[value]),
                        pct=(count / total) if total else 0.0,
                    )
                )

    return rows, warnings


def _require_openpyxl() -> Any:
    try:
        import openpyxl
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit("Missing dependency: run `pip install -r EAs/requirements.txt`") from exc
    return openpyxl, BarChart, Reference, Font, get_column_letter


def _safe_sheet_name(name: str, used: set[str]) -> str:
    safe = "".join("_" if c in r'[]:*?/\\' else c for c in name).strip("'")[:31] or "Sheet"
    candidate = safe
    i = 2
    while candidate in used:
        suffix = f"_{i}"
        candidate = safe[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


def _ordered_params(survivors: list[SurvivorRow], grids_by_set: dict[str, dict[str, ParamGrid]]) -> list[str]:
    seen: set[str] = set()
    params: list[str] = []
    for base_set in sorted({s.base_set for s in survivors}):
        for param in grids_by_set.get(base_set, {}):
            if param not in seen:
                seen.add(param)
                params.append(param)
    return params


def _append_rows(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = cell.font.copy(bold=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = f"A{header_row + 1}"


def _autosize(ws: Any) -> None:
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 40)


def _add_counts_chart(ws: Any, *, title: str, first_row: int, last_row: int, anchor: str) -> None:
    if last_row < first_row:
        return
    _, BarChart, Reference, _, _ = _require_openpyxl()
    chart = BarChart()
    chart.title = title[:250]
    chart.y_axis.title = "Survivors"
    chart.x_axis.title = "Step"
    data = Reference(ws, min_col=4, min_row=first_row - 1, max_row=last_row)
    categories = Reference(ws, min_col=3, min_row=first_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, anchor)


def write_workbook(
    *,
    output: Path,
    survivors: list[SurvivorRow],
    grids_by_set: dict[str, dict[str, ParamGrid]],
    count_rows: list[CountRow],
) -> None:
    openpyxl, _, _, Font, get_column_letter = _require_openpyxl()
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    used_names = {"Survivors"}
    ws = wb.active
    ws.title = "Survivors"

    params = _ordered_params(survivors, grids_by_set)
    survivor_rows = [_survivor_row(survivor, params) for survivor in survivors]
    _append_rows(ws, SURVIVOR_COLUMNS + params, survivor_rows)
    _autosize(ws)

    ws_counts = wb.create_sheet("ValueCounts")
    used_names.add("ValueCounts")
    counts_data = [_count_row(row) for row in sorted(count_rows, key=_count_sort_key)]
    _append_rows(ws_counts, COUNT_COLUMNS, counts_data)
    for cell in ws_counts["F"][1:]:
        cell.number_format = "0.00%"
    _autosize(ws_counts)

    counts_by_set: dict[str, list[CountRow]] = defaultdict(list)
    survivors_by_set: dict[str, list[SurvivorRow]] = defaultdict(list)
    for row in count_rows:
        counts_by_set[row.base_set].append(row)
    for survivor in survivors:
        survivors_by_set[survivor.base_set].append(survivor)

    for base_set in sorted(survivors_by_set):
        sheet_name = _safe_sheet_name(base_set.replace("/", "_").removesuffix(".set"), used_names)
        ws_set = wb.create_sheet(sheet_name)
        ws_set["A1"] = base_set
        ws_set["A1"].font = Font(bold=True)

        set_params = list(grids_by_set.get(base_set, {}))
        matrix_headers = SURVIVOR_COLUMNS + set_params
        ws_set.append([])
        _append_rows(
            ws_set,
            matrix_headers,
            [_survivor_row(survivor, set_params) for survivor in survivors_by_set[base_set]],
        )

        counts_start = ws_set.max_row + 3
        for col, header in enumerate(COUNT_COLUMNS, start=1):
            cell = ws_set.cell(row=counts_start, column=col, value=header)
            cell.font = Font(bold=True)

        row_cursor = counts_start + 1
        chart_row = counts_start
        rows_by_param: dict[str, list[CountRow]] = defaultdict(list)
        for row in counts_by_set[base_set]:
            rows_by_param[row.param].append(row)
        for param in set_params:
            param_rows = sorted(rows_by_param[param], key=_count_sort_key)
            first_row = row_cursor
            for count_row in param_rows:
                ws_set.append(_count_row(count_row))
                ws_set.cell(row=row_cursor, column=6).number_format = "0.00%"
                row_cursor += 1
            last_row = row_cursor - 1
            anchor = f"{get_column_letter(len(matrix_headers) + 3)}{chart_row}"
            _add_counts_chart(ws_set, title=param, first_row=first_row, last_row=last_row, anchor=anchor)
            chart_row += 15

        _autosize(ws_set)

    wb.save(output)


def print_summary(count_rows: list[CountRow], *, low_threshold: int) -> None:
    zero = [row for row in count_rows if row.count == 0]
    low = [row for row in count_rows if 0 < row.count <= low_threshold]
    print(f"Zero-use steps: {len(zero)}")
    for row in zero[:50]:
        print(f"  ZERO {row.base_set} {row.param}={row.step_value}")
    if len(zero) > 50:
        print(f"  ... {len(zero) - 50} more")

    print(f"Low-use steps (<= {low_threshold}): {len(low)}")
    for row in low[:50]:
        print(f"  LOW {row.base_set} {row.param}={row.step_value} count={row.count}")
    if len(low) > 50:
        print(f"  ... {len(low) - 50} more")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate MT5 survivor step-usage Excel report.")
    parser.add_argument("--best-dir", default="", help=f"Defaults to {DEFAULT_BEST_DIR}")
    parser.add_argument(
        "--set-dir",
        default="",
        help="Base optimization .set directory (or set MT5_SET_DIR)",
    )
    parser.add_argument("--out", default=str(DEFAULT_OUTPUT), help="Output .xlsx path")
    parser.add_argument("--terminal", default=str(DEFAULT_TERMINAL), help="MT5 terminal64.exe path")
    parser.add_argument("--mt5-data", default="", help="MT5 data directory override")
    parser.add_argument("--portable", action="store_true", help="Use terminal directory as data directory")
    parser.add_argument("--low-threshold", type=int, default=1, help="Console threshold for low-use steps")
    parser.add_argument("--allow-empty", action="store_true", help="Exit successfully when no survivors exist")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    set_dir = (
        Path(args.set_dir).expanduser().resolve()
        if args.set_dir.strip()
        else resolve_set_dir(required=True)
    )
    set_index = discover_set_files(set_dir)
    if not set_index:
        print(f"No base .set files found under {set_dir}", file=sys.stderr)
        return 1

    best_dir = (
        Path(args.best_dir).expanduser().resolve()
        if args.best_dir
        else DEFAULT_BEST_DIR
    )

    survivors, grids_by_set, load_warnings = load_survivors(
        best_dir=best_dir,
        set_dir=set_dir,
        set_index=set_index,
    )
    for warning in load_warnings:
        print(f"WARNING: {warning}", file=sys.stderr)
    if not survivors:
        print(f"No survivors found in {best_dir / 'best_survivors.csv'}", file=sys.stderr)
        return 0 if args.allow_empty else 1

    count_rows, count_warnings = aggregate_counts(survivors, grids_by_set)
    for warning in count_warnings:
        print(f"WARNING: {warning}", file=sys.stderr)

    output = Path(args.out).expanduser().resolve()
    write_workbook(output=output, survivors=survivors, grids_by_set=grids_by_set, count_rows=count_rows)
    print(f"Wrote {output}")
    print(f"Survivors: {len(survivors)}")
    print_summary(count_rows, low_threshold=args.low_threshold)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
